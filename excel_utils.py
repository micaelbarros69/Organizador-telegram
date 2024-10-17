import openpyxl

def salvar_dados_no_excel_vistoria(dados):
    caminho_planilha = 'vistoria.xlsx'
    
    # Tentar abrir o arquivo ou criar um novo
    try:
        wb = openpyxl.load_workbook(caminho_planilha)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Adicionar cabeçalhos
        sheet.append(['Tipo', 'SGM'])

    # Adicionar os dados extraídos
    sheet.append(dados)
    
    # Salvar a planilha
    wb.save(caminho_planilha)

def salvar_dados_no_excel(dados):
    caminho_planilha = 'retornos.xlsx'
    
    try:
        wb = openpyxl.load_workbook(caminho_planilha)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['ID', 'Projeto', 'Data', 'Local', 'Observações'])

    sheet.append(dados)
    wb.save(caminho_planilha)